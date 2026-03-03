from .constants import MAPPINGS, PRICING_INPUT_DIR, MASTER_PRICING_FILE, ESTIMATOR_TEMPLATE, \
    ESTIMATOR_WORK_SHEET, keyword_map, partner_cost_breakfix_file_map, partner_cost_install_file_map
import pandas as pd
import os, re, shutil, warnings
from openpyxl import load_workbook
from datetime import datetime
from config import settings

support_map = {
    '4hr': '4hr OS',
    'NBD': 'NBD OS',
}
sector_map = {
    'Public': 'Public sector',
    'Commercial': 'Commercial sector',
}


class PriceEstimator:
    def __init__(self, obj):
        self.obj = obj
        self.work_order_initialization()

    def work_order_initialization(self):
        self.sector = sector_map.get(self.obj.sector, '')
        self.support = support_map.get(self.obj.support, '')
        self.country = self.obj.country
        self.term = self.obj.term
        print(f"\nInitialized PriceEstimator with sector: {self.sector}, support: {self.support}, country: {self.country.code}, term: {self.term}")
        
        primary_skus = []
        additional_skus = []
        if hasattr(self.obj, 'primary_sku_items'):
            primary_skus = self.obj.primary_sku_items.all()
        if hasattr(self.obj, 'additional_sku_items'):
            additional_skus = self.obj.additional_sku_items.all()

        self.skus_list = []
        for sku_item in primary_skus:
            self.skus_list.append(sku_item)
        for sku_item in additional_skus:
            self.skus_list.append(sku_item)

        self.pricing_sheet = self.country.pricing_sheet if self.country and hasattr(self.country, 'pricing_sheet') else None
        # print(f"Country pricing sheet name: {self.pricing_sheet}")

    def estimate_price(self):
        pricing_file_path = os.path.join(PRICING_INPUT_DIR, MASTER_PRICING_FILE)
        input_data = []
        print(f'Pricing file path: {pricing_file_path} for Country: {self.country.code} and sheet: {self.pricing_sheet}')
        for sku in self.skus_list:
            original_sku = sku.sku.master_sku_name
            mapped_sku = sku.sku.partner_sku_name
            qty = sku.quantity
            print(f"\n=== Calculating for SKU: {mapped_sku} original: {original_sku}, Qty: {qty} ===")
            data = self.calculate_service_rates(
                excel_path=pricing_file_path,
                sheet_name=self.pricing_sheet,
                sku=mapped_sku,
                country=self.country,
                sector=self.sector,
                term=self.term,
                support=self.support,
                run_all="yes",  # Use "yes" to run all calculations
                sku_item_obj=sku.sku
            )

            if not data:
                print(f"⚠ No data returned for SKU: {original_sku}")
                continue
            data["qty"] = qty
            data["SKU"] = original_sku  # Overwrite mapped name with original name
            input_data.append(data)

        print("\n✅ Completed calculations for all SKUs.")

        work_order_number = self.obj.work_order_number
        wo_estimator_file_name = f'{work_order_number}-{get_timestamp()}.xlsx'
        output_file_path = os.path.join(settings.MEDIA_ROOT, wo_estimator_file_name)
        self.populate_estimator(input_data, output_file=output_file_path)
        return wo_estimator_file_name
    
    def calculate_service_rates(self, excel_path, sheet_name, sku, country, sector, term, support, run_all, sku_item_obj):
        try:
            xls = pd.ExcelFile(excel_path)
            df = xls.parse(sheet_name, header=None)

            header_df = df.iloc[2:6, :].T
            header_df.columns = ['Country', 'Sector', 'Term', 'Support']
            header_df = header_df.dropna().map(lambda x: str(x).strip())

            matched = header_df[
                (header_df['Country'] == country.code) &
                (header_df['Sector'] == sector) &
                (header_df['Term'] == str(term)) &
                (header_df['Support'] == support)
            ]
            if matched.empty:
                raise ValueError(f"No matching rate column found in pricing sheet {sheet_name} for the given parameters.")
            rate_col = matched.index[0]

            sku_row = df[df[0] == sku.strip()].index[0]
            if run_all == "yes":
                keyword_values = {}
                sku_normalized = " ".join(sku.strip().split())
                if sku_normalized in ["GDC-Base SKU", "GDC-Base V2"]:
                    keyword_map["pmo"]["type"] = "PMO"
                else:
                    keyword_map["pmo"]["type"] = "Trade Compliance"

                for keyword, config in keyword_map.items():
                    try:
                        # --- special case: rfp ---
                        if keyword == "rfp":
                            keyword_values[keyword] = round(self.get_service_price_monthly(df, rate_col, sku_row), 2)
                            continue
                        if config.get("external_lookup"):
                            if keyword == "partner cost-break fix":
                                support_key = support.strip().lower()
                                pc_file = partner_cost_breakfix_file_map.get(support_key)
                                pc_file = os.path.join(PRICING_INPUT_DIR, pc_file)
                                # pc_file = prepare_file_path(pc_file)
                                if not pc_file:
                                    raise ValueError(f"No partner cost-break fix file configured for support='{support}'")

                                if support_key == "4hr os":
                                    mapped_sku = sku_item_obj.partner_mapping.master3_breakfix_4hr
                                elif support_key == "nbd os":
                                    mapped_sku = sku_item_obj.partner_mapping.master3_breakfix_nbd
                                else:
                                    raise ValueError(f"Unsupported support level: {support}")

                            elif keyword == "partner cost-call center":
                                pc_file = os.path.join(PRICING_INPUT_DIR, 'Master 3 -Park Place cost for First Call.xlsx')
                                # pc_file = prepare_file_path(pc_file)
                                mapped_sku = sku_item_obj.partner_mapping.master3_first_call
                            elif keyword == "partner cost-install":
                                sector_key = sector.strip().lower()
                                pc_file = partner_cost_install_file_map.get(sector_key)
                                pc_file = os.path.join(PRICING_INPUT_DIR, pc_file)
                                # pc_file = prepare_file_path(pc_file)
                                if not pc_file:
                                    raise ValueError(f"No partner cost-install file configured for sector '{sector}'")
                                
                                pc_df = pd.read_excel(pc_file, header=1)
                                if country.code == "US":
                                    lookup_country = "USA"
                                else:
                                    lookup_country = country.name
                                pc_df['Country'] = pc_df['Country'].astype(str).str.strip().str.lower()
                                row = pc_df[pc_df['Country'].str.contains(lookup_country.lower(), na=False)]
                                if row.empty:
                                    raise ValueError(f"Country '{lookup_country}' not found in Partner Cost Install sheet.")
                                # mapped_sku = sku_map.get(sku.strip(), sku.strip())
                                mapped_sku = sku_item_obj.partner_mapping.master4_deploy_per_install
                                if mapped_sku not in pc_df.columns:
                                    base_name = os.path.basename(pc_file)
                                    raise ValueError(f"Mapped SKU '{mapped_sku}' not found in {base_name} columns.")
                                cost_val = parse_cost(row[mapped_sku].values[0])
                                keyword_values[keyword] = round(cost_val, 2)
                                continue
                            elif keyword == "wwt internal - integration cost":
                                pc_file = os.path.join(PRICING_INPUT_DIR, "Master 2 - WWT-PriceList.xlsx")
                                # pc_file = prepare_file_path(pc_file)
                                mapped_sku_group = sku_item_obj.partner_mapping.master2_price_list
                                # mapped_sku_group = sku_map.get(sku.strip(), sku.strip())

                                wwt_df = pd.read_excel(pc_file, header=None)
                                # Find the row containing the mapped_sku_group in column B (index 1)
                                sku_row_idx = wwt_df[wwt_df[1].astype(str).str.strip() == mapped_sku_group].index[0]
                                # Search for 'NAIC' below this row in column B
                                wwt_region = country.wwt_region.strip()
                                naic_row_idx = wwt_df[(wwt_df.index > sku_row_idx) & (wwt_df[1].astype(str).str.strip() == wwt_region)].index[0]
                                cost_val = parse_cost(wwt_df.at[naic_row_idx, 2])  # Column C is index 2
                                keyword_values[keyword] = round(cost_val, 2)
                                continue

                            else:
                                raise ValueError(f"Unknown external_lookup keyword: {keyword}")

                            pc_df = pd.read_excel(pc_file)
                            # lookup_country = country_map.get(country, country)
                            lookup_country = country.name.strip()
                            pc_df['Country'] = pc_df['Country'].astype(str).str.strip()
                            if lookup_country not in pc_df['Country'].values:
                                raise ValueError(f"{lookup_country} not found in Partner Cost sheet.")
                            # mapped_sku = sku_map.get(sku.strip(), sku.strip())
                            if mapped_sku not in pc_df.columns:
                                base_name = os.path.basename(pc_file)
                                raise ValueError(f"Mapped SKU '{mapped_sku}' not found in {base_name} columns.")
                            matched_row = pc_df[pc_df['Country'] == lookup_country]
                            # matched_row = pc_df[pc_df['Country'].str.contains(lookup_country.lower(), na=False)]
                            col_value = matched_row[mapped_sku]
                            if col_value.values:
                                keyword_values[keyword] = round(parse_cost(col_value.values[0]), 2)
                            else:
                                keyword_values[keyword] = 0.0
                            continue

                        service = config["service"]
                        sum_all = config["sum_all"]
                        type_desc_map, service_row, next_service_row = self.get_all_service_types(df, service, sku_row)
                        # print(f'Type-Description map for service "{service}": {type_desc_map}')
                        if sum_all == "yes":
                            result = self.get_service_cost(df, rate_col, service_row, next_service_row)
                        else:
                            type_name = config.get("type", "")
                            descriptions = []

                            if not type_name:
                                continue
                            if "descriptions" in config and config["descriptions"]:
                                descriptions = config["descriptions"]
                            elif len(type_desc_map.get(type_name, [])) == 1:
                                descriptions = type_desc_map[type_name]
                            elif keyword == "delivery":
                                if sector.lower() == "public sector":
                                    descriptions = ["Secure Logistics and Transportation"]
                                elif sector.lower() == "commercial":
                                    descriptions = ["Commercial Transportation"]

                            if not descriptions:
                                continue

                            result = self.get_service_cost(df, rate_col, service_row, next_service_row, type_name, descriptions)

                        keyword_values[keyword] = result
                    except Exception as e:
                        print(f"{keyword}: Error - {e}")

                parsed_row = {"SKU": sku, "qty": 1}
                parsed_row.update(keyword_values)
                return parsed_row
            
        except Exception as e:
            print(f"Error: {e}")
            return {}
    
    def get_service_price_monthly(self, df, rate_col, sku_row):
        """
        Finds 'Service Price (monthly) -->' within the current SKU section and 
        returns the value in the given rate_col.
        """
        # End of current SKU block: next populated SKU in col 0
        next_sku_rows = df[(df.index > sku_row) & df[0].notna()].index
        next_sku_row = next_sku_rows[0] if len(next_sku_rows) > 0 else len(df)

        def norm(x):
            return " ".join(str(x).replace("\x00", " ").split()).lower()

        target = "service price (monthly)"
        label_row = None
        for r in range(sku_row, len(df)):
            if any(target in norm(v) for v in df.iloc[r, :].tolist()):
                label_row = r
                break

        if label_row is None or rate_col >= df.shape[1]:
            return 0.0

        return parse_cost(df.iat[label_row, rate_col])
    
    def get_all_service_types(self, df, service: str, sku_row: int) -> dict:
        service_row = df[(df[1] == service) & (df.index > sku_row)].index[0]
        next_service_rows = df[(df.index > service_row) & (df[1].notna())].index
        next_service_row = next_service_rows[0] if len(next_service_rows) > 0 else len(df)
        # print(f"Service '{service}' found at row {service_row}, next service row at {next_service_row}")

        type_desc_map = {}
        current_type = None

        for i in range(service_row, next_service_row):
            type_val = str(df.at[i, 2]).replace('\n', ' ').replace('\r', ' ').strip() if pd.notna(df.at[i, 2]) else None
            desc_val = df.at[i, 3]

            if pd.notna(type_val):
                current_type = " ".join(str(type_val).split())
                type_desc_map[current_type] = []

            if pd.notna(desc_val) and current_type:
                type_desc_map[current_type].append(str(desc_val).strip())

            # print(f"Row {i}: Type='{type_val}', Description='{desc_val}'")
        # print(f"Type-Description Map: {type_desc_map}")
        return type_desc_map, service_row, next_service_row
    
    def get_service_cost(self, df, rate_col, service_row, next_service_row, type_name=None, descriptions=None):
        cost_block = df.iloc[service_row:next_service_row, [2, 3, rate_col]]
        cost_block.columns = ['Type', 'Description', 'Cost']
        cost_block = cost_block.dropna(subset=['Cost'])

        if type_name and descriptions:
            filtered = cost_block[
                (cost_block['Type'].ffill().apply(lambda x: " ".join(str(x).split())) == " ".join(type_name.strip().split())) &
                (cost_block['Description'].isin(descriptions))
            ].copy()
        else:
            filtered = cost_block[cost_block['Cost'].notna()].copy()

        filtered['ParsedCost'] = filtered['Cost'].apply(parse_cost)
        return round(filtered['ParsedCost'].sum(), 2)
    
    def populate_estimator(self, input_data_list, output_file='Filled_Estimator.xlsm'):
        # output_file = os.path.join(settings.MEDIA_ROOT, output_file)
        estimator_temp_path = os.path.join(PRICING_INPUT_DIR, ESTIMATOR_TEMPLATE)
        # --- Copy template ---
        print("\n📦 Input Data:")
        for idx, record in enumerate(input_data_list, 1):
            print(f"Product {idx}:")
            for key, value in record.items():
                print(f"  {key}: {value}")

        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        shutil.copyfile(estimator_temp_path, output_file)
        wb = load_workbook(output_file, keep_vba=True, data_only=False)
        ws = wb[ESTIMATOR_WORK_SHEET]

        print(f"Created estimator file: {output_file}")

        # --- Detect sections ---
        placeholders = {
            "main": {},
            "monthly": {},
            "partner": {},
            "internal": {}
        }

        for row in range(1, ws.max_row + 1):
            cell_val = ws[f"A{row}"].value
            if not isinstance(cell_val, str):
                continue

            cell_val_lower = cell_val.strip().lower()
            if cell_val_lower.startswith("monthlyproduct"):
                placeholders["monthly"][cell_val_lower] = row
            elif cell_val_lower.startswith("partnerproduct"):
                placeholders["partner"][cell_val_lower] = row
            elif cell_val_lower.startswith("internalproduct"):
                placeholders["internal"][cell_val_lower] = row
            elif cell_val_lower.startswith("product"):
                placeholders["main"][cell_val_lower] = row

        # --- Apply input data ---
        for idx, data in enumerate(input_data_list):
            tag = f"product{idx + 1}".lower()
            monthly_tag = f"monthlyproduct{idx + 1}".lower()

            # Main
            if tag in placeholders["main"]:
                row = placeholders["main"][tag]
                ws[f"A{row}"] = data["SKU"]
                ws[f"B{row}"] = data["qty"]
                ws[f"O{row}"] = data.get("rfp", "0")

                formula = ws[f"C{row}"].value
                formula = formula.replace('0', str(data.get("integration", "0.0")), 1)
                ws[f"C{row}"] = formula
                
                formula = ws[f"D{row}"].value
                formula = formula.replace('0', str(data.get("delivery", "0.0")), 1)
                ws[f"D{row}"] = formula

                formula = ws[f"E{row}"].value
                formula = formula.replace('0', str(data.get("title", "0.0")), 1)
                ws[f"E{row}"] = formula

                formula = ws[f"G{row}"].value
                formula = formula.replace('0', str(data.get("install", "0.0")), 1)
                ws[f"G{row}"] = formula

                formula = ws[f"H{row}"].value
                formula = formula.replace('0', str(data.get("pmo", "0.0")), 1)
                ws[f"H{row}"] = formula
                
                formula = ws[f"I{row}"].value
                formula = formula.replace('0', str(data.get("break/fix", "0.0")), 1)
                ws[f"I{row}"] = formula

                formula = ws[f"J{row}"].value
                formula = formula.replace('0', str(data.get("call center", "0.0")), 1)                
                ws[f"J{row}"] = formula

            # Monthly
            if monthly_tag in placeholders["monthly"]:
                row = placeholders["monthly"][monthly_tag]
                ws[f"A{row}"] = data["SKU"]
                ws[f"B{row}"] = data["qty"]
            
            # Internal
            internal_tag = f"internalproduct{idx + 1}".lower()
            if internal_tag in placeholders["internal"]:
                row = placeholders["internal"][internal_tag]
                ws[f"A{row}"] = data["SKU"]
                ws[f"B{row}"] = data["qty"]
                formula = ws[f"C{row}"].value
                formula = formula.replace('0', str(data.get("wwt internal - integration cost", "0.0")), 1)
                ws[f"C{row}"] = formula
            
            # Partner
            partner_tag = f"partnerproduct{idx + 1}".lower()
            if partner_tag in placeholders["partner"]:
                row = placeholders["partner"][partner_tag]
                ws[f"A{row}"] = data["SKU"]
                ws[f"B{row}"] = data["qty"]
                formula = ws[f"C{row}"].value
                formula = formula.replace('0', str(data.get("partner cost-install", "0.0")), 1)
                ws[f"C{row}"] = formula

                formula = ws[f"D{row}"].value
                formula = formula.replace('0', str(data.get("partner cost-break fix", "0.0")), 1)
                ws[f"D{row}"] = formula

                formula = ws[f"E{row}"].value
                formula = formula.replace('0', str(data.get("partner cost-call center", "0.0")), 1)
                ws[f"E{row}"] = formula

        # --- Clear unused placeholders in column A ---
        max_items = len(input_data_list)
        for section, tags in placeholders.items():
            for tag, row in tags.items():
                match = re.search(r'(\d+)$', tag)
                if match and int(match.group(1)) > max_items:
                    ws[f"A{row}"] = None

        # --- Save ---
        wb.save(output_file)
        print(f"✅ Updated {len(input_data_list)} rows. Output: {output_file}")


## Utility methods
def get_timestamp():
    # return datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
    return datetime.now().strftime("%Y-%m-%d_%H%M")

def norm(v): 
    return str(v).strip()

def parse_cost(val):
    try:
        return float(str(val).replace('$', '').strip())
    except:
        return 0.0